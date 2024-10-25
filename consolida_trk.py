import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import os

usuario_atual = os.getlogin()

arquivos = [f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/02 NORDESTE.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/01 NORTE.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/03 LESTE.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/04 FARMA BRASIL.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/05 ALIMENTAR DIRETO.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/06 ALIMENTAR INDIRETO.xlsx',
            f'C:/Users/{usuario_atual}/Baruel/Baruel - CUSTOMER SERVICE/CUSTOMER SERVICE/01. ACOMPANHAMENTO GERAL/00. PLANILHAS DE ACOMPANHAMENTO/07 CENTRO OESTE.xlsx']

saida = 'C:/relato/arquivo_consolidado_teste.xlsx'
def consolida_trk(arquivos, saida):
    dfs = [pd.read_excel(arquivos, sheet_name='Base') for arquivos in arquivos]

    print('---------------------')
    print(' - ARQUIVOS BAIXADOS COM SUCESSO')

    consolida_df = pd.concat(dfs, ignore_index=True)

    print(' - CONSOLIDANDO ARQUIVOS')
    print(' - FORMATANDO AS COLUNAS COM DATA')

    consolida_df.to_excel(saida, index=False)



    # Abrir o arquivo Excel existente ou criar um novo
    try:
        wb = load_workbook(saida)
        ws = wb.active
        # Deletar todos os dados na planilha
        ws.delete_rows(1, ws.max_row)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    # Escrever o DataFrame consolidado na planilha
    for r in dataframe_to_rows(consolida_df, index=False, header=True):
        ws.append(r)

    wb = load_workbook(saida)


    # Definir o autor nos metadados do arquivo
    wb.properties.creator = 'Jean Lino'
    aba = wb['Sheet1']
    aba.title = 'Base'
    # Salvar novamente o arquivo com o autor atualizado
    wb.save(saida)

    print(' - ARQUIVO CONSOLIDADO SALVO EM: ' + saida)
    print('---------- ;-) -----------')

def organiza_colunas_data(saida):
    wb = load_workbook(saida)
    ws = wb.active  # Seleciona a primeira planilha

    # Estilo para a célula no formato de data
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")



    colunas = (3, 19, 20, 22, 24, 25, 26, 27, 37, 47)
    print('Formatando datas, por favor aguarde')
    for i in colunas: # itera sobre a tupla colunas

        # Definindo a coluna desejada (por exemplo, coluna 3)
        coluna_desejada = i  # C    oluna C

        # Itera pelas células da coluna definida (exemplo: da linha 2 até a linha 10)
        for row in ws.iter_rows(min_row=2, min_col=coluna_desejada, max_row=20000, max_col=coluna_desejada):
            for cell in row:
                if isinstance(cell.value, datetime):
                    # Se o valor já é uma data, apenas aplica o estilo
                    cell.style = date_style
                else:
                    try:
                        # Converte o valor da célula em data, se for uma string
                        date_value = datetime.strptime(cell.value, '%d/%m/%Y')  # Ajuste o formato se necessário
                        cell.value = date_value
                        cell.style = date_style
                    except (ValueError, TypeError):
                        print(f"Valor na célula {cell.coordinate} não é uma data válida ou não é uma string.")


        # Salvar as alterações no arquivo Excel
        wb.save(saida)


consolida_trk(arquivos, saida)
organiza_colunas_data(saida)
